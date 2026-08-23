import os
import json
import csv
import math
from datetime import datetime, date
from .config import get_active_paths

def format_time(seconds):
    try:
        if seconds is None or not isinstance(seconds, (int, float)) or not math.isfinite(seconds):
            sec_val = 0
        else:
            sec_val = max(0, int(seconds))
    except Exception:
        sec_val = 0

    mins, secs = divmod(sec_val, 60)
    hrs, mins = divmod(mins, 60)
    return f"{hrs:02d}h {mins:02d}m {secs:02d}s" if hrs > 0 else f"{mins:02d}m {secs:02d}s"

def format_short_time(seconds):
    try:
        if seconds is None or not isinstance(seconds, (int, float)) or not math.isfinite(seconds):
            sec_val = 0
        else:
            sec_val = max(0, int(seconds))
    except Exception:
        sec_val = 0

    mins, secs = divmod(sec_val, 60)
    hrs, mins = divmod(mins, 60)
    return f"{hrs:02d}:{mins:02d}:{secs:02d}" if hrs > 0 else f"{mins:02d}:{secs:02d}"

def normalize_task_name(task_name):
    if not task_name or not isinstance(task_name, str):
        return "DEEP_WORK"
    cleaned = "_".join(task_name.strip().split()).upper()
    return cleaned if cleaned else "DEEP_WORK"

def load_all_sessions():
    _, data_file, _ = get_active_paths()
    if not os.path.exists(data_file):
        return []
    sessions = []
    try:
        with open(data_file, "r", encoding="utf-8") as f:
            for line in f:
                line_str = line.strip()
                if line_str:
                    try:
                        record = json.loads(line_str)
                        if isinstance(record, dict):
                            # Sanitize fields
                            focus_sec = record.get("focus_seconds")
                            if not isinstance(focus_sec, (int, float)) or not math.isfinite(focus_sec) or focus_sec < 0:
                                record["focus_seconds"] = 0.0
                            else:
                                record["focus_seconds"] = float(focus_sec)

                            break_sec = record.get("break_seconds")
                            if not isinstance(break_sec, (int, float)) or not math.isfinite(break_sec) or break_sec < 0:
                                record["break_seconds"] = 0.0
                            else:
                                record["break_seconds"] = float(break_sec)

                            if not isinstance(record.get("date"), str) or not record["date"].strip():
                                record["date"] = date.today().strftime("%Y-%m-%d")

                            if not isinstance(record.get("start_time"), str):
                                record["start_time"] = "00:00:00"

                            if not isinstance(record.get("end_time"), str):
                                record["end_time"] = "00:00:00"

                            record["task"] = normalize_task_name(record.get("task"))

                            sessions.append(record)
                    except json.JSONDecodeError:
                        continue
    except Exception:
        pass
    return sessions

def overwrite_all_sessions(sessions):
    _, data_file, _ = get_active_paths()
    try:
        with open(data_file, "w", encoding="utf-8") as f:
            for s in sessions:
                if isinstance(s, dict):
                    f.write(json.dumps(s) + "\n")
        sync_markdown_report()
    except Exception as e:
        print(f"\033[1;31mError saving session store: {e}\033[0m\n")

def save_session(start_dt, end_dt, focus_seconds, break_seconds, task_name="DEEP_WORK"):
    _, data_file, _ = get_active_paths()
    record = {
        "date": start_dt.strftime("%Y-%m-%d"),
        "start_time": start_dt.strftime("%H:%M:%S"),
        "end_time": end_dt.strftime("%H:%M:%S"),
        "focus_seconds": round(max(0.0, float(focus_seconds)), 2),
        "break_seconds": round(max(0.0, float(break_seconds)), 2),
        "task": normalize_task_name(task_name)
    }
    try:
        with open(data_file, "a", encoding="utf-8") as f:
            f.write(json.dumps(record) + "\n")
        sync_markdown_report()
    except Exception as e:
        print(f"\033[1;31mError appending session: {e}\033[0m\n")

def sync_markdown_report():
    sessions = load_all_sessions()
    _, _, md_file = get_active_paths()
    try:
        if not sessions:
            with open(md_file, "w", encoding="utf-8") as f:
                f.write("# ⚡ Flowmodoro Deep Work Journal\n\n*No focus sessions recorded yet.*\n")
            return

        total_focus = sum(s.get("focus_seconds", 0.0) for s in sessions)
        total_breaks = sum(s.get("break_seconds", 0.0) for s in sessions)
        by_date = {}
        by_task = {}
        for s in sessions:
            d = s.get("date", date.today().strftime("%Y-%m-%d"))
            t = s.get("task", "Deep Work")
            f_sec = s.get("focus_seconds", 0.0)
            by_date.setdefault(d, []).append(s)
            by_task[t] = by_task.get(t, 0.0) + f_sec

        today_str = date.today().strftime("%Y-%m-%d")
        today_sessions = by_date.get(today_str, [])
        today_focus = sum(s.get("focus_seconds", 0.0) for s in today_sessions)

        with open(md_file, "w", encoding="utf-8") as f:
            f.write("# ⚡ Flowmodoro Deep Work Journal\n\n")
            f.write(f"> *Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*\n\n")
            f.write("## 📊 Overview Metrics\n\n")
            f.write("| Metric | Value |\n| :--- | :--- |\n")
            f.write(f"| **Today's Focus** | `{format_time(today_focus)}` ({len(today_sessions)} sessions) |\n")
            f.write(f"| **All-Time Focus** | `{format_time(total_focus)}` |\n")
            f.write(f"| **Total Rest Earned** | `{format_time(total_breaks)}` |\n")
            f.write(f"| **Total Completed Cycles** | `{len(sessions)}` |\n")
            f.write(f"| **Active Days** | `{len(by_date)}` |\n\n")

            if by_task and total_focus > 0:
                f.write("## 🎯 Focus Share by Topic\n\n| Objective / Task | Focus Time | Share (%) |\n| :--- | :--- | :--- |\n")
                sorted_tasks = sorted(by_task.items(), key=lambda x: x[1], reverse=True)
                for t_name, t_sec in sorted_tasks:
                    pct = (t_sec / total_focus) * 100
                    f.write(f"| {t_name} | `{format_time(t_sec)}` | `{pct:.1f}%` |\n")
                f.write("\n")

            f.write("## 📝 Session Logs\n\n| Date | Task / Topic | Start | End | Focus | Earned Break |\n| :--- | :--- | :--- | :--- | :--- | :--- |\n")
            for s in reversed(sessions):
                date_val = s.get('date', '')
                task_val = s.get('task', 'Deep Work')
                start_val = s.get('start_time', '00:00:00')
                end_val = s.get('end_time', '00:00:00')
                f_sec = s.get('focus_seconds', 0.0)
                b_sec = s.get('break_seconds', 0.0)
                f.write(f"| {date_val} | {task_val} | {start_val} | {end_val} | `{format_short_time(f_sec)}` | `{format_short_time(b_sec)}` |\n")

    except Exception as e:
        print(f"\033[1;31mError generating markdown report: {e}\033[0m\n")

def delete_last_session():
    _, _, md_file = get_active_paths()
    sessions = load_all_sessions()
    if not sessions:
        print("\n\033[1;31mNo sessions available to remove.\033[0m\n")
        return
    last = sessions[-1]
    print(f"\nMost recent session: [{last.get('date')} {last.get('start_time')}-{last.get('end_time')}] '{last.get('task')}' ({format_short_time(last.get('focus_seconds'))})")
    try:
        choice = input("Are you sure you want to remove this session? [y/N]: ").strip().lower()
        if choice == 'y':
            removed = sessions.pop()
            overwrite_all_sessions(sessions)
            print(f"\033[1;32m✓ Removed session '{removed.get('task')}' and resynced {md_file}\033[0m\n")
    except (KeyboardInterrupt, EOFError):
        print("\nOperation canceled.\n")

def parse_mass_selection(input_str, sessions):
    if not input_str or not isinstance(input_str, str):
        return set()
    cleaned = input_str.strip().lower()
    total = len(sessions)
    if cleaned in ("all", "all sessions", "*"):
        return set(range(total))

    selected_indices = set()
    tokens = [t.strip() for t in cleaned.split(",") if t.strip()]
    for token in tokens:
        if "-" in token:
            parts = token.split("-")
            if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                start_num = int(parts[0])
                end_num = int(parts[1])
                for num in range(min(start_num, end_num), max(start_num, end_num) + 1):
                    if 1 <= num <= total:
                        selected_indices.add(num - 1)
        elif token.isdigit():
            num = int(token)
            if 1 <= num <= total:
                selected_indices.add(num - 1)
        else:
            for idx, s in enumerate(sessions):
                t_val = str(s.get("task", "")).lower()
                d_val = str(s.get("date", "")).lower()
                if token in t_val or token in d_val:
                    selected_indices.add(idx)

    return selected_indices


def interactive_delete_session():
    _, _, md_file = get_active_paths()
    sessions = load_all_sessions()
    if not sessions:
        print("\n\033[1;31mNo sessions recorded yet.\033[0m\n")
        return
    os.system('cls' if os.name == 'nt' else 'clear')
    print("=" * 65 + "\n                🗑️  DELETE / MANAGE SESSIONS\n" + "=" * 65)
    display_limit = min(20, len(sessions))
    start_idx = len(sessions) - display_limit
    print(f"\nShowing last {display_limit} sessions (newest at bottom):\n")
    print("  #   | Date       | Time Window     | Task                 | Focus\n  ----+------------+-----------------+----------------------+----------")
    for i in range(start_idx, len(sessions)):
        s = sessions[i]
        d_val = s.get('date', 'N/A')
        st_val = s.get('start_time', '00:00:00')
        et_val = s.get('end_time', '00:00:00')
        t_val = str(s.get('task', 'Deep Work'))[:20]
        f_val = format_short_time(s.get('focus_seconds', 0))
        print(f"  {i+1:<3} | {d_val} | {st_val} - {et_val:<8} | {t_val:<20} | {f_val}")
    print("=" * 65)
    print("\033[2mMass selection examples: '1, 3, 5' | '1-5' | '1-3, 7' | 'all' | 'taskname'\033[0m")
    try:
        choice = input(f"\nEnter session #(s), range, task name, or 'all' (or 'q' to cancel): ").strip().lower()
        if choice not in ('q', ''):
            to_delete = parse_mass_selection(choice, sessions)
            if not to_delete:
                print("\033[1;31mNo matching sessions found to delete.\033[0m\n")
                return

            sorted_indices = sorted(list(to_delete))
            print(f"\n\033[1;33mSelected {len(sorted_indices)} session(s) to delete:\033[0m")
            for idx in sorted_indices[:10]:
                s = sessions[idx]
                print(f"  • #{idx+1}: [{s.get('date')} {s.get('start_time')}] '{s.get('task')}' ({format_short_time(s.get('focus_seconds'))})")
            if len(sorted_indices) > 10:
                print(f"  ... and {len(sorted_indices) - 10} more session(s).")

            confirm = input(f"\nAre you sure you want to mass delete these {len(sorted_indices)} session(s)? [y/N]: ").strip().lower()
            if confirm == 'y':
                remaining = [s for idx, s in enumerate(sessions) if idx not in to_delete]
                overwrite_all_sessions(remaining)
                print(f"\033[1;32m✓ Mass deleted {len(to_delete)} session(s) and resynced {md_file}\033[0m\n")
    except (KeyboardInterrupt, EOFError):
        print("\nOperation canceled.\n")

def delete_by_task(task_query):
    if not task_query or not isinstance(task_query, str):
        print("\033[1;31mError: Please specify a task name query to delete (e.g. flowmodoro --delete-task 'test').\033[0m\n")
        return
    sessions = load_all_sessions()
    if not sessions:
        print("\033[1;31mNo sessions available to delete.\033[0m\n")
        return
    query = task_query.strip().lower()
    matching = [idx for idx, s in enumerate(sessions) if query in str(s.get("task", "")).lower()]
    if not matching:
        print(f"\033[1;31mNo sessions matching '{task_query}' were found.\033[0m\n")
        return
    print(f"\033[1;33mFound {len(matching)} session(s) matching task '{task_query}':\033[0m")
    for idx in matching[:10]:
        s = sessions[idx]
        print(f"  • #{idx+1}: [{s.get('date')} {s.get('start_time')}] '{s.get('task')}' ({format_short_time(s.get('focus_seconds'))})")
    if len(matching) > 10:
        print(f"  ... and {len(matching) - 10} more session(s).")
    try:
        confirm = input(f"\nAre you sure you want to mass delete all {len(matching)} session(s)? [y/N]: ").strip().lower()
        if confirm == 'y':
            remaining = [s for idx, s in enumerate(sessions) if idx not in matching]
            overwrite_all_sessions(remaining)
            print(f"\033[1;32m✓ Mass deleted {len(matching)} session(s) matching '{task_query}'.\033[0m\n")
    except (KeyboardInterrupt, EOFError):
        print("\nOperation canceled.\n")

def delete_all_sessions():
    sessions = load_all_sessions()
    if not sessions:
        print("\033[1;31mNo session history to clear.\033[0m\n")
        return
    print(f"\033[1;31mWARNING: This will permanently delete ALL {len(sessions)} recorded focus session(s)!\033[0m")
    try:
        confirm = input("Type 'DELETE ALL' to confirm: ").strip()
        if confirm == "DELETE ALL":
            overwrite_all_sessions([])
            print(f"\033[1;32m✓ All session history has been cleared.\033[0m\n")
        else:
            print("Operation canceled.\n")
    except (KeyboardInterrupt, EOFError):
        print("\nOperation canceled.\n")


def export_to_excel(dest_path, sessions):

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.chart import DoughnutChart, BarChart, LineChart, Reference
    except ImportError:
        print("\033[1;31mError: openpyxl module is required for .xlsx export. Run 'pip install openpyxl' first.\033[0m\n")
        return

    try:
        wb = openpyxl.Workbook()

        # Sheet 1: Session Logs
        ws_logs = wb.active
        ws_logs.title = "Session Logs"
        ws_logs.views.sheetView[0].showGridLines = True

        headers = ["Date", "Task / Topic", "Start Time", "End Time", "Focus (Sec)", "Focus (Min)", "Focus (Hours)", "Break (Sec)", "Break (Min)"]
        ws_logs.append(headers)

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_idx in range(1, len(headers) + 1):
            cell = ws_logs.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for s in sessions:
            f_sec = s.get("focus_seconds", 0.0)
            b_sec = s.get("break_seconds", 0.0)
            ws_logs.append([
                s.get("date", ""),
                s.get("task", "Deep Work"),
                s.get("start_time", ""),
                s.get("end_time", ""),
                round(f_sec, 2),
                round(f_sec / 60.0, 2),
                round(f_sec / 3600.0, 2),
                round(b_sec, 2),
                round(b_sec / 60.0, 2)
            ])

        for col in ws_logs.columns:
            max_l = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws_logs.column_dimensions[col_letter].width = max(max_l + 4, 12)

        # Sheet 2: Visual Dashboard
        ws_dash = wb.create_sheet(title="Visual Dashboard")
        ws_dash.views.sheetView[0].showGridLines = True

        by_task = {}
        by_date = {}
        for s in sessions:
            t = s.get("task", "Deep Work")
            d = s.get("date", "")
            f_hrs = s.get("focus_seconds", 0.0) / 3600.0
            b_mins = s.get("break_seconds", 0.0) / 60.0
            by_task[t] = by_task.get(t, 0.0) + f_hrs
            if d not in by_date:
                by_date[d] = {"focus_hrs": 0.0, "break_mins": 0.0}
            by_date[d]["focus_hrs"] += f_hrs
            by_date[d]["break_mins"] += b_mins

        # Task Table
        ws_dash.cell(row=1, column=1, value="Task / Topic").font = Font(bold=True)
        ws_dash.cell(row=1, column=2, value="Focus Hours").font = Font(bold=True)
        row_idx = 2
        for t_name, t_hrs in sorted(by_task.items(), key=lambda x: x[1], reverse=True):
            ws_dash.cell(row=row_idx, column=1, value=t_name)
            ws_dash.cell(row=row_idx, column=2, value=round(t_hrs, 2))
            row_idx += 1
        task_end_row = row_idx - 1

        # Date Table
        ws_dash.cell(row=1, column=4, value="Date").font = Font(bold=True)
        ws_dash.cell(row=1, column=5, value="Focus Hours").font = Font(bold=True)
        ws_dash.cell(row=1, column=6, value="Break Mins").font = Font(bold=True)
        row_d_idx = 2
        for d_val, d_data in sorted(by_date.items()):
            ws_dash.cell(row=row_d_idx, column=4, value=d_val)
            ws_dash.cell(row=row_d_idx, column=5, value=round(d_data["focus_hrs"], 2))
            ws_dash.cell(row=row_d_idx, column=6, value=round(d_data["break_mins"], 2))
            row_d_idx += 1
        date_end_row = row_d_idx - 1

        # Chart 1: Donut Chart
        if task_end_row >= 2:
            chart1 = DoughnutChart()
            chart1.title = "1. Deep Work Share by Task"
            chart1.style = 10
            chart1.width = 14
            chart1.height = 10
            labels1 = Reference(ws_dash, min_col=1, min_row=2, max_row=task_end_row)
            data1 = Reference(ws_dash, min_col=2, min_row=1, max_row=task_end_row)
            chart1.add_data(data1, titles_from_data=True)
            chart1.set_categories(labels1)
            ws_dash.add_chart(chart1, "H2")

        # Chart 2: Clustered Bar Chart
        if date_end_row >= 2:
            chart2 = BarChart()
            chart2.type = "col"
            chart2.style = 11
            chart2.title = "2. Daily Focus Hours Progress"
            chart2.y_axis.title = "Hours"
            chart2.x_axis.title = "Date"
            chart2.width = 16
            chart2.height = 10
            labels2 = Reference(ws_dash, min_col=4, min_row=2, max_row=date_end_row)
            data2 = Reference(ws_dash, min_col=5, min_row=1, max_row=date_end_row)
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(labels2)
            ws_dash.add_chart(chart2, "H18")

        # Chart 3: Line Chart
        if len(sessions) >= 1:
            chart3 = LineChart()
            chart3.title = "3. Session Flow Endurance Trend"
            chart3.style = 13
            chart3.y_axis.title = "Focus Minutes"
            chart3.x_axis.title = "Session #"
            chart3.width = 16
            chart3.height = 10
            data3 = Reference(ws_logs, min_col=6, min_row=1, max_row=len(sessions) + 1)
            chart3.add_data(data3, titles_from_data=True)
            ws_dash.add_chart(chart3, "R2")

        # Chart 4: Stacked Bar Chart
        if date_end_row >= 2:
            chart4 = BarChart()
            chart4.type = "col"
            chart4.grouping = "stacked"
            chart4.overlap = 100
            chart4.title = "4. Focus vs. Earned Rest (Minutes)"
            chart4.y_axis.title = "Minutes"
            chart4.x_axis.title = "Date"
            chart4.width = 16
            chart4.height = 10
            labels4 = Reference(ws_dash, min_col=4, min_row=2, max_row=date_end_row)
            data4 = Reference(ws_dash, min_col=5, min_row=1, max_col=6, max_row=date_end_row)
            chart4.add_data(data4, titles_from_data=True)
            chart4.set_categories(labels4)
            ws_dash.add_chart(chart4, "R18")

        wb.save(dest_path)
        print(f"\033[1;32m✓ Exported {len(sessions)} session(s) with 4 pre-built Excel charts to:\033[0m\n  📂 {dest_path}\n")
    except Exception as e:
        print(f"\033[1;31mExcel export failed: {e}\033[0m\n")


def export_data(destination_file):
    """Exports session logs to CSV, JSON, or XLSX format."""
    if not destination_file or not isinstance(destination_file, str) or "\0" in destination_file:
        print("\033[1;31mExport failed: Invalid destination path specified.\033[0m\n")
        return

    cleaned = destination_file.strip()
    if not cleaned:
        print("\033[1;31mExport failed: Destination file path cannot be empty.\033[0m\n")
        return

    sessions = load_all_sessions()
    if not sessions:
        print("\n\033[1;31mNo session logs found to export.\033[0m\n")
        return

    try:
        dest_path = os.path.abspath(os.path.expanduser(cleaned))
        if os.path.isdir(dest_path):
            print(f"\033[1;31mExport failed: Path '{dest_path}' is a directory, not a file.\033[0m\n")
            return

        parent_dir = os.path.dirname(dest_path)
        if parent_dir:
            os.makedirs(parent_dir, exist_ok=True)

        ext = os.path.splitext(dest_path)[1].lower()

        if ext == ".xlsx":
            export_to_excel(dest_path, sessions)
        elif ext == ".json":
            with open(dest_path, "w", encoding="utf-8") as f:
                json.dump(sessions, f, indent=2)
            print(f"\033[1;32m✓ Exported {len(sessions)} session(s) successfully to:\033[0m\n  📂 {dest_path}\n")
        else:  # Default to CSV
            if not ext.endswith(".csv"):
                dest_path += ".csv"
            with open(dest_path, "w", newline="", encoding="utf-8") as f:
                fields = ["date", "task", "start_time", "end_time", "focus_seconds", "focus_minutes", "focus_hours", "break_seconds", "break_minutes"]
                writer = csv.DictWriter(f, fieldnames=fields)
                writer.writeheader()
                for s in sessions:
                    f_sec = s.get("focus_seconds", 0.0)
                    b_sec = s.get("break_seconds", 0.0)
                    writer.writerow({
                        "date": s.get("date"),
                        "task": s.get("task", "Deep Work"),
                        "start_time": s.get("start_time"),
                        "end_time": s.get("end_time"),
                        "focus_seconds": round(f_sec, 2),
                        "focus_minutes": round(f_sec / 60.0, 2),
                        "focus_hours": round(f_sec / 3600.0, 2),
                        "break_seconds": round(b_sec, 2),
                        "break_minutes": round(b_sec / 60.0, 2)
                    })
            print(f"\033[1;32m✓ Exported {len(sessions)} session(s) successfully to:\033[0m\n  📂 {dest_path}\n")

    except Exception as e:
        print(f"\033[1;31mExport failed: {e}\033[0m\n")


